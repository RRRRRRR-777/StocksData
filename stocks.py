########################################

# * import設定
import os
import sys
import pandas as pd
from bs4 import BeautifulSoup
import re
import requests
import time
import numpy as np
from typing_extensions import Protocol
import json 
import random
import yahoo_fin.stock_info as si
import html5lib
from datetime import datetime as dt
from datetime import timedelta, timezone
import pathlib
import glob
import pprint
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import mplfinance as mpf
import shutil
from IPython.core.display import Image
from openpyxl.workbook.workbook import Workbook
import openpyxl
from statistics import mean
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
import math

JST = timezone(timedelta(hours=+9), 'JST')
dt_now = dt.now(JST)
date = dt_now.strftime('%y%m%d')
start = time.time()

print ('import completed')

# データベース作成

base_url = 'https://finviz.com/screener.ashx?v=151&f=cap_smallover,geo_usa,ind_stocksonly&C='+str(list(range(71))).replace(' ','').replace('[','').replace(']','')
df = pd.DataFrame()
cnt = 1
while 1:
    url = base_url+'&r='+str(cnt)
    time.sleep(1.0)
    site = requests.get(url, headers={'User-Agent': 'Custom'}, timeout=3.5)
    data = BeautifulSoup(site.text,'html.parser')
    da = data.find_all("tr", align="center")
    tables = re.findall('<td class="table-top.*</td>', str(da[0]))
    names = [re.sub('.*>(.+)</td>.*', r'\1', s) for s in tables]
    cntns = re.findall('<a class="screener-link.*</a>', str(da[0]))
    contents = [[re.sub('<a .*">(.+)</a>.*', r'\1', s).replace('</span>','').replace('&amp;','&') for s in re.findall('<a .*?</a>', i)] for i in cntns]
    df1 = pd.DataFrame(contents, columns = names)
    print("\r now reading -->>  " +df1['Ticker'][0]+'('+str(cnt)+")  ---" ,end="")
    df = pd.concat([df, df1], axis=0)
    if(len(df1)!=20): break
    cnt+=20

df = df.replace('-', np.nan)
conv_nums = ['Market Cap','Outstanding','Float','Avg Volume']
conv_float = ['P/E','Fwd P/E','PEG','P/S','P/B','P/C','P/FCF','EPS','Short Ratio','Curr R','Quick R','LTDebt/Eq',
             'Debt/Eq','Beta','ATR','RSI','Recom','Rel Volume','Price','Target Price']
conv_pct = ['Dividend','Payout Ratio','EPS this Y','EPS next Y','EPS past 5Y','EPS next 5Y','Sales past 5Y','EPS Q/Q','Sales Q/Q',
           'Insider Own','Inst Own','Inst Trans','Float Short','ROA','ROE','ROI','Gross M','Oper M','Profit M',
           'Perf Week','Perf Month','Perf Quart','Perf Half','Perf Year','Perf YTD',
           'Volatility W','Volatility M','SMA20','SMA50','SMA200','50D High','50D Low','52W High','52W Low',
           'from Open','Gap','Change']
conv_date = ['IPO Date']
volume = ['Volume']

for c in df.columns.to_list():
  if (c in conv_nums):
    df[c] = [float(str(s).translate(str.maketrans({'K':'E3','M':'E6','B':'E9','T':'E12'}))) for s in df[c]]
  elif (c in conv_float):
    df[c] = [float(s) for s in df[c]]
  elif (c in conv_pct):
    df[c] = [float(str(s).replace('%',''))/100 for s in df[c]]
  elif (c in conv_date):
    df[c] = [dt.strptime(str(s), '%m/%d/%Y') if re.match(r'\d+/\d+/\d+', str(s)) else s for s in df[c]]
  elif (c in volume):
    df[c] = [int(str(s).replace(',','')) for s in df[c]]
   
df['PrevClose'] = df['Price'] / (1+df['Change'])
df = df.set_index('Ticker')



def set_param(df_i, name_ta, param_ta, cat_name=''):
 if(cat_name!=''):
   df_i[cat_name] = ''
 for i in range(len(param_ta)):
   base_url = 'https://finviz.com/screener.ashx?v=521&'+param_ta[i]
   df_p = pd.DataFrame()
   cnt = 1
   print("\r now reading -->>  " +name_ta[i]+"  ---" ,end="")
   while 1:
     url = base_url+'&r='+str(cnt)
     time.sleep(1)
     site = requests.get(url, headers={'User-Agent': 'Custom'}, timeout=3.5)
     data = BeautifulSoup(site.text,'html.parser')
     da = data.find_all("td", align="center")
     tmp1 = [str(s) for s in da if 'cssbody=[tabchrtbdy]' in str(s)]
     tmp2 = [[s for s in re.findall('<td title=.*?</small></td>', s)] for s in tmp1]
     flat = [x for row in tmp2 for x in row]
     tickers = [re.sub('.*">(.+)</a>.*', r'\1', s) for s in flat]
     df1 = pd.DataFrame({ "Ticker" : tickers })
     if(len(df1)>0):
       print("\r now reading -->>  " +name_ta[i]+' : ' +df1['Ticker'][0]+'('+str(cnt)+")  ---" ,end="")
     df_p = pd.concat([df_p, df1], axis=0)
     if(len(df1)<500): break
     cnt+=500
   if(cat_name!=''):
     df_i.loc[df_p['Ticker'].to_list(),cat_name] = [s+' '+name_ta[i] if s!='' else name_ta[i] for s in df_i.loc[df_p['Ticker'].to_list(),cat_name].to_list()]
   else:
     df_i[name_ta[i]] = 0
     df_i.loc[df_p['Ticker'].to_list(),name_ta[i]] = 1
 return 



# 追加パラメータ
url = 'https://finviz.com/screener.ashx?v=111&ft=4'
site = requests.get(url, headers={'User-Agent': 'Custom'}, timeout=3.5)
data = BeautifulSoup(site.text,'html.parser')
value_list = [[s['value'] for s in dat.find_all('option') if 'value' in str(s)] for dat in data.find_all("select")]
name_list = [[s.text.strip() for s in dat.find_all('option') if 'value' in str(s)] for dat in data.find_all("select")]




def get_param(df, num, pre_str, first_cut, last_cut ,cat_name=''):
 if first_cut:
   value_list[num].pop(0)
   name_list[num].pop(0)
 if last_cut:
   value_list[num].pop(-1)
   name_list[num].pop(-1)
 ret_param = [pre_str+s.replace('v=111','').replace('ft=4','').replace('screener.ashx?','').replace('&','') for s in value_list[num]]
 ret_name = name_list[num]
 set_param(df, ret_name, ret_param, cat_name)
 return


get_param(df, 3, '', True, False)  # Signal
get_param(df, 4, 'f=exch_', True, True, 'Exchange')  # Exchange
get_param(df, 5, 'f=idx_', True, False, 'Index')  # Index
get_param(df, 42, 'f=sh_opt_', True, True)  # Option/Short
get_param(df, 58, 'f=ta_candlestick_', True, True, 'Candlestick') # Candlestickrp


df.drop_duplicates(inplace=True)


# RS CALC 
df_rs = df.copy()
df_rs = df_rs.fillna({'Perf Week': 0, 'Perf Month': 0, 'Perf Quart': 0, 'Perf Half': 0, 'Perf Year': 0})
# RS 計算でETFの除きたい場合は以下のコメントを外す
# df_rs = df_rs[df_rs['Industry']!='Exchange Traded Fund']
df_rs['price_mid'] = df_rs['Perf Quart']
df_rs.loc[(df_rs['price_mid'] == 0), 'price_mid'] = df_rs['Perf Month']
df_rs.loc[(df_rs['price_mid'] == 0), 'price_mid'] = df_rs['Perf Week']
df_rs['price_last'] = df_rs['Perf Year']
df_rs.loc[(df_rs['price_last'] == 0), 'price_last'] = df_rs['Perf Half']
df_rs.loc[(df_rs['price_last'] == 0), 'price_last'] = df_rs['Perf Quart']
df_rs.loc[(df_rs['price_last'] == 0), 'price_last'] = df_rs['Perf Month']
df_rs.loc[(df_rs['price_last'] == 0), 'price_last'] = df_rs['Perf Week']
df_rs['POS_NOW'] = 100*(1+df_rs['price_last'])
df_rs['POS_MID'] = df_rs['POS_NOW'] / (1+df_rs['price_mid'])
df_rs['RS_Sort'] = df_rs['POS_MID']+(df_rs['POS_NOW']-df_rs['POS_MID'])*2
df_rs.sort_values('RS_Sort', ascending=True, inplace=True)
df_rs['RS']=[s/(len(df_rs)+1)*100 for s in range(1,len(df_rs)+1)]
df['RS'] = 0
df.loc[df_rs.index, 'RS'] = df_rs['RS']
numbers = df['No.'].to_numpy().tolist()
for i in range(len(numbers)):
    df.iloc[i,109] = '{:.2f}'.format(df.iloc[i,109])



try:
  p = glob.glob('input/stock_data_*.csv', recursive=True)[0] 
  os.remove(p)
except:
  pass

out_file = 'input/stock_data_'+date+'.csv'
df.to_csv(out_file, encoding='utf_8_sig')
elapsed_time = time.time() - start
print("\nelapsed_time: {0}".format(elapsed_time) + "[sec]")
print("Save:",out_file)


########################################


#  *  RS70未満の銘柄と株価＄20未満の銘柄削除


p = glob.glob('input/stock_data_*.csv', recursive=True)[0]
df_all_stocks = pd.read_csv(p)
df_RS = df_all_stocks.query('RS >= 70 & Price >= 20').rename(columns={'52W High':'High_52','52W Low':'Low_52','Avg Volume':'Avg_Volume'})
tickers = df_RS['Ticker'].to_numpy().tolist()

print("RS70未満の銘柄と株価$20未満の銘柄削除 completed")


#  *  アナリストデータの追加


cnt = 1
df1=pd.DataFrame()
for i in range(len(tickers)):
  try:
      analysts = si.get_analysts_info(tickers[i])
      print("\r now reading -->>  " +"$"+tickers[i]+'('+str(cnt)+")  --- ",end="")
      cnt+=1
      try:
        EPS = analysts['EPS Trend'].iloc[0]
        CQ_EPS,NQ_EPS,CY_EPS,NY_EPS = EPS.iloc[1],EPS.iloc[2],EPS.iloc[3],EPS.iloc[4]
        df1.loc[tickers[i],'今期EPS']='$%.2f'%(CQ_EPS)
        df1.loc[tickers[i],'来期EPS']='$%.2f'%(NQ_EPS)
        df1.loc[tickers[i],'今年度EPS']='$%.2f'%(CY_EPS)
        df1.loc[tickers[i],'来年度EPS']='$%.2f'%(NY_EPS)
        Revenue = analysts['Revenue Estimate'].iloc[1]
        CQ_Rev,NQ_Rev,CY_Rev,NY_Rev = Revenue.iloc[1],Revenue.iloc[2],Revenue.iloc[3],Revenue.iloc[4]
        df1.loc[tickers[i],'今期Rev']= CQ_Rev
        df1.loc[tickers[i],'来期Rev']= NQ_Rev
        df1.loc[tickers[i],'今年度Rev']= CY_Rev
        df1.loc[tickers[i],'来年度Rev']= NY_Rev
      except KeyError:
        print("\r Ecception KeyError by EPS or Rvenue -->>  " +"$"+tickers[i]+'('+str(cnt)+")  ---")
        cnt+=1
  except ValueError:
      print("\r Ecception ValueError by analysts -->>  " +"$"+tickers[i]+'('+str(cnt)+")  ---")
      cnt+=1


#  *  ATH


tickers = df_RS['Ticker'].to_numpy().tolist()
cnt = 1
df2=pd.DataFrame()
for i in range(len(tickers)):
  try:
    ATH = si.get_data(tickers[i])['high'].max()
    df2.loc[tickers[i],'ATH'] = float('%.2f'%(ATH))
    print("\r now reading -->>  " +"$"+tickers[i]+'('+str(cnt)+")  --- ",end="")
  except:
    print('\r error -->>'+"$"+tickers[i]+'('+str(cnt)+")")
  cnt+=1


#  *  U/D Volume Raito 作成


cnt = 1
df3=pd.DataFrame()
for t in range(len(tickers)):
  try:
    up = 0
    down = 0
    data = si.get_data(tickers[t]).dropna().tail(50).sort_index(axis=0, ascending=False)
    close = data['adjclose']
    volume = data['volume']
    for i in range(0,49):
      if close[i] > close[i+1]:
        up = up + volume[i]
      else:
        down = down + volume[i]
    UDVR = float('%.2f'%(up/down))
    df3.loc[tickers[t],'UDVR'] = UDVR
    print("\r now reading -->>  " +"$"+tickers[t]+'('+str(cnt)+")  --- ",end="")
    cnt+=1
  except IndexError:
    print("\r Ecception IndexError -->>  " +"$"+tickers[i]+'('+str(cnt)+")  ---")
    cnt+=1


#  *  df_RS, AnalystsData, ATH を 合わせてCSV出力


# df1 アナリストデータ
df1['Ticker'] = df1.index
# df2 ATH
df2['Ticker'] = df2.index
# df3 U/D Volume Ratio
df3['Ticker'] = df3.index
df = pd.merge(df1,df2, on='Ticker') 
df = pd.merge(df, df3)
df_RS_newdata = df_RS.merge(df, on='Ticker') 
print(df_RS_newdata)
try:
    p = glob.glob('input/RS_newdata_*.csv', recursive=True)[0] 
    os.remove(p)
except:
    pass
out_file = 'input/RS_newdata_'+date+'.csv'
df_RS_newdata.to_csv(out_file, encoding='utf_8_sig')


########################################


#  *  CAN-SLIM条件をクリアしている銘柄にカウントを与え、順位を振り分ける


# df作成
p = glob.glob('input/RS_newdata_*.csv', recursive=True)[0]
df = pd.read_csv(p,index_col=0)
df = df.rename(columns={'EPS Q/Q':'EPS_Q', 'Sales Q/Q':'Sales_Q','EPS past 5Y':'EPS_past_5Y',
                        'Sales past 5Y':'Sales_past_5Y', 'Market Cap':'Market_Cap',
                        'Insider Own':'Insider_Own', 'Inst Trans':'Inst_Trans'})
df['CANSLIM_cnt'] = 0
# 背景色設定
def color_background_lightgreen(val):
    color = 'lightgreen'  
    return 'background-color: %s' % color

# <---C---> EPS Q/Q, Sales Q/Q  (今期EPS,Salesの昨対がともにプラス)
C = df.query('EPS_Q > 0.25 & Sales_Q > 0.25')['CANSLIM_cnt']+1
# <---A---> EPS Past5year, Salse Past5year がともにプラス
#           ROEが17％以上
A = df.query('EPS_past_5Y >= 0 & Sales_past_5Y >= 0')['CANSLIM_cnt'] + 1
A2 = df.query('ROE >= 0.17')['CANSLIM_cnt'] + 1
# <---N---> 52High or ATH
N = df.query('High_52 >= 0 or Price >= ATH')['CANSLIM_cnt']+1
# <---S---> InsiderOwn (MarkertCap >= 10000000000, 0.01 | MarketCap < 10000000000, 0.03)
# UPDR >= 1
S = df.query('Market_Cap >= 10000000000 & Insider_Own >= 0.01')['CANSLIM_cnt']+1
S2 = df.query('Market_Cap < 10000000000 & Insider_Own >= 0.03')['CANSLIM_cnt']+1
S3 = df.query('UDVR >= 1')['CANSLIM_cnt']+1
# <---L---> すべての銘柄がRS70以上をクリアしているのでRSを表示するだけ
# df['RS']
# <---I---> InstTransがプラス
I = df.query('Inst_Trans >= 0')['CANSLIM_cnt']+1
# 株価が50SMAの上にあり50SMAが200SMAの上にある
SMA = df.query('SMA200>SMA50>0')['CANSLIM_cnt']+1
# 株価が出来高を伴って上昇している
VOL = df.query('Volume > Avg_Volume * 1.4 & Change > 0')['CANSLIM_cnt']+1
# dfにCANSLIM_CNTの追加
CANSLIM = pd.concat([C, A, A2, N, S, S2, S3, I, SMA, VOL], axis=1).fillna(0)
CANSLIM_CNT = CANSLIM.sum(axis=1).dropna(0)
CANSLIM_CNT.name='CANSLIM_CNT'
df = pd.concat([df, CANSLIM_CNT], axis=1).drop('CANSLIM_cnt', axis=1)
# dfをCANSLIM＿CNT順位付けしてTicker,CANSLIM_CNT,RANKのみ表示
df = df[['Ticker', 'Market_Cap', 'CANSLIM_CNT']]
df = pd.concat([df, df['CANSLIM_CNT'].rank(ascending=False, method='min')], axis=1)
df.columns = ['Ticker', 'Market_Cap', 'CANSLIM_CNT', 'rank']
df = df.sort_values(by='rank', ascending=True).dropna()
# 各順位を時価総額順で順位付け
Mkt_df=pd.DataFrame()
rank_list = df['rank'].unique().tolist()
for i in range(0, len(rank_list), 1):
    rank_num = rank_list[i]    
    num = df[df['rank'] == rank_num]
    MktCap = pd.concat([df[df['rank'] == rank_num], num['Market_Cap'].rank(ascending=False, method='min')], axis=1)
    MktCap.columns = ['Ticker', 'Market_Cap', 'CANSLIM_CNT', 'rank', 'Mkt_RANK']
    MktCap = MktCap.sort_values(by='Mkt_RANK', ascending=True)
    Mkt_df = pd.concat([Mkt_df, MktCap])
RANK_df = Mkt_df.copy().reindex(columns=['Ticker','Market_Cap','Mkt_RANK','CANSLIM_CNT','rank'])
p = glob.glob('input/RS_newdata_*.csv', recursive=True)[0]
RS_newdata_df = pd.read_csv(p,index_col=0)
RS_newdata_df = pd.merge(RANK_df, RS_newdata_df, on='Ticker').drop(['Market_Cap', 'Mkt_RANK', 'No.'], axis=1).rename(columns={})
RS_newdata_df = RS_newdata_df.rename(columns={'EPS Q/Q':'EPS_Q', 'Sales Q/Q':'Sales_Q','EPS past 5Y':'EPS_past_5Y',
                        'Sales past 5Y':'Sales_past_5Y', 'Market Cap':'Market_Cap',
                        'Insider Own':'Insider_Own', 'Inst Trans':'Inst_Trans'})
# rankの重複をなくす
tickers = RS_newdata_df['Ticker'].to_numpy().tolist()
RS_newdata_df['RANK'] = 0
RS_newdata_df = RS_newdata_df.set_index('Ticker')
for i in range(0, len(tickers), 1):
    RS_newdata_df.loc[tickers[i], 'RANK'] = i+1
# 不要な列を削除
RS_newdata_df = RS_newdata_df.drop(['P/E', 'Fwd P/E', 'PEG', 'P/S', 'P/B',
       'P/C', 'P/FCF', 'Dividend', 'Payout Ratio', 
       'Short Ratio', 'Curr R', 'Quick R',
       'LTDebt/Eq', 'Debt/Eq', 'Gross M', 'Oper M', 'Profit M',
        'Beta', 'ATR', 'Volatility W', 'Volatility M',  'from Open', 'Rel Volume',
        'Top Gainers', 'Top Losers', 'New High',
       'New Low', 'Most Volatile', 'Most Active', 'Unusual Volume',
       'Overbought', 'Oversold', 'Downgrades', 'Upgrades',
       'Earnings Before', 'Earnings After', 'Recent Insider Buying',
       'Recent Insider Selling', 'Major News', 'Horizontal S/R',
       'TL Resistance', 'TL Support', 'Wedge Up', 'Wedge Down',
       'Triangle Ascending', 'Triangle Descending', 'Wedge', 'Channel Up',
       'Channel Down', 'Channel', 'Double Top', 'Double Bottom',
       'Multiple Top', 'Multiple Bottom', 'Head & Shoulders',
       'Head & Shoulders Inverse', 'Exchange', 'Index', 'Optionable',
       'Shortable', 'Candlestick', 'Gap', 'rank'], axis=1)
# 列の入れ替え
RS_newdata_df = RS_newdata_df.reindex(columns=['Ticker','Company', 'Sector', 'Industry',
       'Country', 'RS','RANK', 'CANSLIM_CNT', 'Market_Cap','Avg_Volume', 'Price', 'Change',
       'Volume', 'UDVR', 'EPS', 'EPS this Y', 'EPS next Y','EPS_past_5Y', 'EPS next 5Y',
       'Sales_past_5Y', 'EPS_Q', 'Sales_Q','Outstanding', 'Float', 'Insider_Own',
       'Insider Trans', 'Inst Own','Inst_Trans', 'Float Short', 'ROA', 'ROE', 'ROI',
       'Perf Week','Perf Month', 'Perf Quart', 'Perf Half', 'Perf Year', 'Perf YTD',
       'SMA20', 'SMA50', 'SMA200', '50D High', '50D Low', 'High_52','Low_52', 'ATH',
       'RSI', 'Recom', 'Earnings', 'Target Price', 'IPO Date', 'PrevClose', '今期EPS',
       '来期EPS', '今年度EPS', '来年度EPS', '今期Rev', '来期Rev','今年度Rev', '来年度Rev'])
RS_newdata_df['Ticker'] = RS_newdata_df.index
# CSV出力
outfile = 'input/CANSLIM_'+date+'.csv'
try:
  p = glob.glob('input/CANSLIM_*.csv', recursive=True)[0]    
  os.remove(p)
except:
  pass 
RS_newdata_df.head(100).to_csv(outfile, index=False, encoding='utf_8_sig')
print("Saved : ",outfile)


#  *  CAN-SLIM条件にマッチしているものに色を付ける


p = glob.glob('input/CANSLIM_*.csv', recursive=True)[0]
df = pd.read_csv(p)

def func(s):
    style_list = []
    for i in range(0, len(df.columns.values), 1):
       style_list.append('')

    if float(s['EPS_Q']) >= 0.25  and float(s['Sales_Q']) >= 0.25:
       style_list[20] = 'background-color:palegreen'
       style_list[21] = 'background-color:palegreen'
    if float(s['EPS_past_5Y']) >= 0  and float(s['Sales_past_5Y']) >= 0:
       style_list[17] = 'background-color:palegreen'
       style_list[19] = 'background-color:palegreen'
    if float(s['ROE']) >= 0.17:
       style_list[30] = 'background-color:palegreen'
    if float(s['Market_Cap']) >= 10000000000 and float(s['Insider_Own']) >= 0.1:
       style_list[24] = 'background-color:palegreen'
    if float(s['Market_Cap']) < 10000000000 and float(s['Insider_Own']) >= 0.3:
       style_list[24] = 'background-color:palegreen'
    if float(s['UDVR']) >= 1:
       style_list[13] = 'background-color:palegreen'
    if float(s['Inst_Trans']) > 0:
      style_list[27] = 'background-color:palegreen'
    if float(s['SMA200']) > float(s['SMA50']) > float(0):
      style_list[39] = 'background-color:palegreen'
      style_list[40] = 'background-color:palegreen'
    if float(s['Volume']) >= float(s['Avg_Volume'] * 1.4) and float(s['Change']) > 0:
      style_list[12] = 'background-color:palegreen'
    return style_list

CANSLIM_df = df.style.apply(func, axis=1)
# Excelに出力
outfile = 'input/CANSLIM_'+date+'.xlsx'
try:
    p = glob.glob('input/CANSLIM_*.xlsx', recursive=True)[0]    
    os.remove(p)
except:
    pass
CANSLIM_df.to_excel(outfile, sheet_name='CAN_SLIM_Screening', index=False, encoding='utf_8_sig')
print('saved : ExcelFile')


########################################


csvfile = glob.glob('input/CANSLIM_*.csv', recursive=True)[0]
excelfile = glob.glob('input/CANSLIM_*.xlsx', recursive=True)[0] 
outfile = 'input/CANSLIM_'+date+'.xlsx'
photosfile='input/Pictures_'+date+'/'

wb = openpyxl.load_workbook(filename=excelfile)
ws = wb['CAN_SLIM_Screening']
ws2 = wb.create_sheet(title='ChartPhotos')


# ChartPhotos作成
try:
  p = glob.glob('input/Pictures_*/', recursive=True)[0]
  shutil.rmtree(p)
except:
  pass
os.makedirs(photosfile, exist_ok=True)

df = pd.read_csv(csvfile, index_col=0)
tickers = df.index[:100].to_numpy().tolist()
cnt=1
for i in range(0, len(tickers), 1):
  # 株価を260日分取得し、60日分の最初の日にも200日移動平均線を描画
  df = si.get_data(tickers[i]).tail(260)
  # 移動平均線の定義付け
  df['SMA50'] = df['adjclose'].rolling(50).mean().round(2)
  df['SMA200'] = df['adjclose'].rolling(200).mean().round(2)
  df['volumeSMA60'] = df['volume'].rolling(60).mean().round(0)
  df['EMA21'] = df['adjclose'].ewm(span=21).mean()
  # dfを60日分に変更
  df=df.tail(60).dropna(how='all', axis=1)
  # 作図
  adds=[]
  if 'SMA50' in df.columns:
      adds.append(mpf.make_addplot(df['SMA50'], color='g', width=1.5, alpha=0.5))
  if 'SMA200' in df.columns:
      adds.append(mpf.make_addplot(df['SMA200'], color='r', width=1.5, alpha=0.5))
  if 'EMA21' in df.columns:
      adds.append(mpf.make_addplot(df['EMA21'], color='b', width=1.5, alpha=0.5))
  if 'volumeSMA60' in df.columns:
      adds.append(mpf.make_addplot(df['volumeSMA60'], color='b', width=1.5, alpha=0.8, panel='lower'))
  else:
    mpf.plot(df, type='candle', volume=True, style='yahoo', figsize=(7.2, 4.5), title='Daily Chart $'+tickers[i], savefig=photosfile+tickers[i]) 
    print('none addplot -->>  '+"$"+tickers[i]+'('+str(cnt)+")  --- ") 
    cnt+=1
  mpf.plot(df, type='candle', volume=True, style='yahoo', figsize=(7.2, 4.5), title='Daily Chart $'+tickers[i], addplot=adds, savefig=photosfile+tickers[i])
  print("\r now reading -->>  " +"$"+tickers[i]+'('+str(cnt)+")  --- ",end="")
  cnt+=1
print('\r'+photosfile)

# Sheet['ChartPhotos']の作成
data = pd.read_csv(csvfile, index_col=0)
tickers = data.index[:100].to_numpy().tolist()

ws2.column_dimensions['A'].width = 400
for i in range(0, len(tickers), 1):
    img_dir = ('input/Pictures_'+date+'/'+tickers[i]+'.png')
    img_to_excel = openpyxl.drawing.image.Image(img_dir)
    ws2.add_image(img_to_excel, 'A'+str(i+1))
    ws2.row_dimensions[i+1].height = 400
    cell = '#ChartPhotos!$A${}'.format(i+1)
    name = '{}'.format(tickers[i])
    ws.cell(i+2,1).value = '=HYPERLINK("{0}","{1}")'.format(cell, name)

# Table設定
def num2alpha(num):
    if num<=26:
        return chr(64+num)
    elif num%26==0:
        return num2alpha(num//26-1)+chr(90)
    else:
        return num2alpha(num//26)+chr(64+num%26)
max_rows = ws.max_row
max_columns = num2alpha(ws.max_column)
cell_range = 'A1:%s%d'%(max_columns, max_rows)

table = Table(displayName='CANSLIM', ref=cell_range, )
style = TableStyleInfo(name='TableStyleLight1', showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)
# 列の固定と文字サイズの変更
ws.freeze_panes = 'A2'
font = Font(size=14)

for row in ws:
    for cell in row:
        ws[cell.coordinate].font = font

# エクセルのセル幅を自動調整
max_row = ws.max_row
alpha = [num2alpha(n) for n in range(1, ws.max_column+1)]
for c in range(len(alpha)):
    column = alpha[c]
    l = []
    m = 0
    for i in range(0, max_row, 1):
      l.append(len(re.sub('\.','',str(ws[column][i].value))))
      if column=='A':
        ws.column_dimensions[column].width = 14
      else:
        if column=='B' or column=='C' or column=='D':
          pass
        else:
          z = [z for z in range(1, len(l)) if l[z]>=math.floor(mean(l))*2]
          y = [ l[z[y]] for y in range(len(z))]
          [l.remove(y[x]) for x in range(len(y))]
        m = max(l)
        ws.column_dimensions[column].width = (m*1.2)+7
    print('\r '+column, end='')
#別名保存
outfile = 'input/CANSLIM_'+date+'.xlsx'
wb.save(outfile)
print('\r '+outfile)

# 使用したファイルを全て削除する
try:
    p = glob.glob('input/stock_data_*.csv', recursive=True)[0]
    os.remove(p)
except:
    pass
try:
    p = glob.glob('input/RS_newdata_*.csv', recursive=True)[0]
    os.remove(p)
except:
    pass
try:
    p = glob.glob('input/CANSLIM_*.csv', recursive=True)[0]
    os.remove(p)
except:
    pass
try:
    p = glob.glob('input/Pictures_*/', recursive=True)[0]
    shutil.rmtree(p)
except:
    pass